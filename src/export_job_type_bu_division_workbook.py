from __future__ import annotations

import base64
import csv
import json
import os
import sys
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Attachment, Disposition, FileContent, FileName, FileType, Mail

from client import ServiceTitanClient

PROJECT = Path('/workspace/apps/lex-servicetitan-reporting')
OUT_DIR = PROJECT / 'data' / 'exports'
STAMP = datetime.now(timezone.utc).strftime('%Y-%m-%d_%H%M%S')

MAX_ATTEMPTS = 8
POLL_DELAY = 40
POLL_TRIES = 10
POLL_INTERVAL = 10
SLEEP_BETWEEN = 12
BLOCK_TOKENS = ('5.7.606', 'banned sending IP', 'Access denied')


def load_env_file(path: str, override: bool = False) -> None:
    p = Path(path)
    if not p.exists():
        return
    for raw in p.read_text().splitlines():
        line = raw.strip()
        if not line or line.startswith('#') or '=' not in line:
            continue
        k, v = line.split('=', 1)
        k = k.strip()
        v = v.strip().strip('"').strip("'")
        if override or k not in os.environ:
            os.environ[k] = v


def paginate(client: ServiceTitanClient, path: str, params: dict[str, Any] | None = None, page_size: int = 500) -> tuple[list[dict[str, Any]], int]:
    rows: list[dict[str, Any]] = []
    page = 1
    pages = 0
    base_params = dict(params or {})
    while True:
        req_params = dict(base_params)
        req_params.update({'page': page, 'pageSize': page_size})
        r = client.get(path, req_params)
        if r.status_code >= 400:
            raise RuntimeError(f'{path} page={page} failed: {r.status_code} {r.text[:500]}')
        body = r.json()
        data = body.get('data', [])
        rows.extend(data)
        pages += 1
        if not body.get('hasMore'):
            return rows, pages
        page += 1


def flatten(row: dict[str, Any], prefix: str = '') -> dict[str, Any]:
    out: dict[str, Any] = {}
    for k, v in row.items():
        key = f'{prefix}{k}'
        if isinstance(v, dict):
            out.update(flatten(v, f'{key}.'))
        elif isinstance(v, list):
            out[key] = json.dumps(v, default=str)
        else:
            out[key] = v
    return out


def is_active(row: dict[str, Any]) -> bool:
    if 'active' in row:
        return bool(row.get('active'))
    if 'isActive' in row:
        return bool(row.get('isActive'))
    return True


def active_label(row: dict[str, Any]) -> Any:
    if 'active' in row:
        return row.get('active')
    if 'isActive' in row:
        return row.get('isActive')
    return True


def sorted_columns(rows: list[dict[str, Any]], preferred: list[str]) -> list[str]:
    all_cols = sorted({k for r in rows for k in r.keys()})
    cols = [c for c in preferred if c in all_cols]
    cols.extend([c for c in all_cols if c not in cols])
    return cols


def write_sheet(wb: Workbook, title: str, rows: list[dict[str, Any]], preferred: list[str]) -> None:
    ws = wb.create_sheet(title)
    flat = [flatten(r) for r in rows]
    cols = sorted_columns(flat, preferred)
    if not cols:
        cols = ['note']
        flat = [{'note': 'No rows returned'}]
    ws.append(cols)
    header_fill = PatternFill('solid', fgColor='1A3A5C')
    header_font = Font(color='FFFFFF', bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    for row in flat:
        ws.append([row.get(c) for c in cols])
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    for idx, col in enumerate(cols, 1):
        max_len = len(str(col))
        for row_idx in range(2, min(ws.max_row, 200) + 1):
            val = ws.cell(row_idx, idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 55)


def derive_divisions(business_units: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen: dict[str, dict[str, Any]] = {}
    for bu in business_units:
        div = bu.get('division') or bu.get('businessUnitDivision') or bu.get('divisionInfo')
        if not isinstance(div, dict):
            continue
        div_id = div.get('id') or div.get('divisionId') or div.get('name') or json.dumps(div, sort_keys=True)
        key = str(div_id)
        item = seen.setdefault(key, {
            'division.id': div.get('id') or div.get('divisionId'),
            'division.name': div.get('name') or div.get('divisionName'),
            'activeBusinessUnitCount': 0,
            'activeBusinessUnitIds': [],
            'activeBusinessUnitNames': [],
            'rawDivision': json.dumps(div, default=str, sort_keys=True),
        })
        item['activeBusinessUnitCount'] += 1
        item['activeBusinessUnitIds'].append(bu.get('id'))
        item['activeBusinessUnitNames'].append(bu.get('name'))
    out = []
    for item in seen.values():
        item['activeBusinessUnitIds'] = ', '.join(str(x) for x in item['activeBusinessUnitIds'] if x is not None)
        item['activeBusinessUnitNames'] = ', '.join(str(x) for x in item['activeBusinessUnitNames'] if x)
        out.append(item)
    out.sort(key=lambda x: (str(x.get('division.name') or ''), str(x.get('division.id') or '')))
    return out


def lookup_status(api_key: str, to_email: str, subject: str):
    q = f'to_email="{to_email}" AND subject="{subject}"'
    try:
        r = requests.get('https://api.sendgrid.com/v3/messages', headers={'Authorization': f'Bearer {api_key}'}, params={'limit': 5, 'query': q}, timeout=45)
    except requests.RequestException as e:
        return None, f'lookup exception: {type(e).__name__}', None
    if r.status_code != 200:
        return None, f'lookup http {r.status_code}', None
    msgs = r.json().get('messages', [])
    if not msgs:
        return None, 'no msg yet', None
    m = msgs[0]
    try:
        d = requests.get(f'https://api.sendgrid.com/v3/messages/{m["msg_id"]}', headers={'Authorization': f'Bearer {api_key}'}, timeout=45).json()
    except requests.RequestException as e:
        return None, f'detail exception: {type(e).__name__}', None
    reason = ''
    for ev in d.get('events', []):
        if ev.get('reason'):
            reason = ev['reason']
    return d.get('status'), reason, d.get('outbound_ip')


def send_with_retry(attachment_path: Path, subject: str, html: str, to: str) -> bool:
    api_key = os.environ['SENDGRID_API_KEY']
    from_email = os.environ.get('SENDGRID_FROM_EMAIL', 'alerts@lexairconditioning.com')
    from_name = os.environ.get('SENDGRID_FROM_NAME', 'LEX ServiceTitan')
    recipients = [e.strip() for e in to.split(',') if e.strip()]
    if not recipients:
        raise RuntimeError('No email recipients configured')
    data = base64.b64encode(attachment_path.read_bytes()).decode()
    for attempt in range(1, MAX_ATTEMPTS + 1):
        mail = Mail(from_email=(from_email, from_name), to_emails=recipients, subject=subject, html_content=html)
        mail.attachment = [Attachment(FileContent(data), FileName(attachment_path.name), FileType('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'), Disposition('attachment'))]
        resp = SendGridAPIClient(api_key).send(mail)
        print(f'[sendgrid] attempt {attempt}/{MAX_ATTEMPTS} submitted status={resp.status_code}')
        time.sleep(POLL_DELAY)
        status = reason = ip = None
        for _ in range(POLL_TRIES):
            status, reason, ip = lookup_status(api_key, recipients[0], subject)
            if status in ('delivered', 'bounce', 'dropped', 'not_delivered'):
                break
            time.sleep(POLL_INTERVAL)
        print(f'[sendgrid] status={status} outbound_ip={ip} reason={(reason or "")[:180]}')
        if status == 'delivered':
            return True
        if status is None or any(t in (reason or '') for t in BLOCK_TOKENS):
            time.sleep(SLEEP_BETWEEN)
            continue
        return False
    return False


def main() -> int:
    load_env_file('/workspace/.secrets/hermes.env')
    load_env_file('/workspace/apps/lex-monthly-insights/.env', override=True)
    load_env_file(str(PROJECT / '.env'), override=True)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    client = ServiceTitanClient()

    print('Fetching active job types with pagination...')
    job_types_all, jt_pages = paginate(client, '/jpm/v2/tenant/{tenant}/job-types', {'active': 'True'}, page_size=500)
    job_types = [r for r in job_types_all if is_active(r)]
    print(f'  pages={jt_pages} rows={len(job_types)} active_rows={len(job_types)}')

    print('Fetching active business units with pagination...')
    business_units_all, bu_pages = paginate(client, '/settings/v2/tenant/{tenant}/business-units', {'active': 'True'}, page_size=500)
    business_units = [r for r in business_units_all if is_active(r)]
    print(f'  pages={bu_pages} rows={len(business_units_all)} active_rows={len(business_units)}')

    divisions = derive_divisions(business_units)
    print(f'  derived_divisions={len(divisions)}')

    workbook = OUT_DIR / f'servicetitan_active_job_types_business_units_divisions_{STAMP}.xlsx'
    wb = Workbook()
    summary = wb.active
    summary.title = 'Summary'
    summary_rows = [
        ['Generated UTC', datetime.now(timezone.utc).isoformat(timespec='seconds')],
        ['Tenant ID', client.cfg.tenant_id],
        ['Job type endpoint', '/jpm/v2/tenant/{tenant}/job-types'],
        ['Job type pages fetched', jt_pages],
        ['Active job types', len(job_types)],
        ['Business unit endpoint', '/settings/v2/tenant/{tenant}/business-units'],
        ['Business unit pages fetched', bu_pages],
        ['Active business units', len(business_units)],
        ['Derived divisions', len(divisions)],
        ['Division note', 'Divisions are derived from nested division objects on active business units. If blank, ServiceTitan did not return division metadata on the active BU rows.'],
    ]
    for row in summary_rows:
        summary.append(row)
    summary['A1'].font = Font(bold=True)
    summary.column_dimensions['A'].width = 32
    summary.column_dimensions['B'].width = 90

    write_sheet(wb, 'Active Job Types', job_types, ['id', 'name', 'active', 'isActive', 'summary', 'businessUnitId', 'businessUnit.id', 'businessUnit.name', 'priority', 'duration', 'noCharge', 'soldThreshold'])
    write_sheet(wb, 'Active Business Units', business_units, ['id', 'name', 'officialName', 'active', 'isActive', 'division.id', 'division.name', 'trade.id', 'trade.name', 'businessUnitType', 'email', 'phoneNumber'])
    if divisions:
        write_sheet(wb, 'Divisions', divisions, ['division.id', 'division.name', 'activeBusinessUnitCount', 'activeBusinessUnitIds', 'activeBusinessUnitNames', 'rawDivision'])
    else:
        write_sheet(wb, 'Divisions', [{'note': 'No division objects were returned on active business units.'}], ['note'])
    wb.save(workbook)
    print(f'Workbook: {workbook}')

    # CSV backups for audit/debug without needing Excel.
    for name, rows in [('job_types', job_types), ('business_units', business_units), ('divisions', divisions)]:
        csv_path = OUT_DIR / f'servicetitan_active_{name}_{STAMP}.csv'
        flat = [flatten(r) for r in rows]
        cols = sorted({k for r in flat for k in r.keys()}) or ['note']
        with csv_path.open('w', newline='') as f:
            w = csv.DictWriter(f, fieldnames=cols)
            w.writeheader()
            for r in flat:
                w.writerow(r)
        print(f'CSV: {csv_path} rows={len(flat)}')

    to = os.environ.get('MONTHLY_REPORT_TO', 'ryan@servicestarbrands.com')
    subject = f'LEX ServiceTitan Active Reference Lists | {STAMP} UTC'
    html = f'''
    <div style="font-family:Helvetica,Arial,sans-serif;color:#1F2937;max-width:680px">
      <div style="background:#1a3a5c;color:#fff;padding:18px 22px;border-bottom:3px solid #DAA520">
        <div style="font-size:20px;font-weight:700">LEX ServiceTitan Active Reference Lists</div>
        <div style="font-size:12px;color:#BFD0E5;margin-top:4px">Job Types, Business Units, and Divisions</div>
      </div>
      <div style="padding:22px;font-size:13px;line-height:1.5">
        <p>Attached is the requested Excel workbook pulled live from ServiceTitan.</p>
        <ul>
          <li><b>Active job types:</b> {len(job_types)} across {jt_pages} API page(s)</li>
          <li><b>Active business units:</b> {len(business_units)} across {bu_pages} API page(s)</li>
          <li><b>Divisions found:</b> {len(divisions)} derived from business-unit division metadata</li>
        </ul>
        <p style="color:#6B7280;font-size:11px">Generated {STAMP} UTC.</p>
      </div>
    </div>
    '''
    delivered = send_with_retry(workbook, subject, html, to)
    manifest = OUT_DIR / f'servicetitan_active_reference_lists_{STAMP}.json'
    manifest.write_text(json.dumps({
        'workbook': str(workbook),
        'job_type_pages': jt_pages,
        'active_job_types': len(job_types),
        'business_unit_pages': bu_pages,
        'active_business_units': len(business_units),
        'divisions': len(divisions),
        'to': to,
        'subject': subject,
        'delivered': delivered,
    }, indent=2))
    print(f'Manifest: {manifest}')
    print('DELIVERED' if delivered else 'NOT_DELIVERED')
    return 0 if delivered else 3


if __name__ == '__main__':
    raise SystemExit(main())
